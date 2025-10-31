# xylella_processor.py — versão final Cloud (corrigida)
# Gere os ficheiros Excel, recolhe debug e cria ZIP com summary.
from __future__ import annotations
import os, io, zipfile, re
from pathlib import Path
from typing import List, Dict, Any, Tuple
import importlib
from openpyxl import load_workbook

# ───────────────────────────────────────────────
# Importação do core_xylella
# ───────────────────────────────────────────────
_CORE_MODULE_NAME = "core_xylella"
core = importlib.import_module(_CORE_MODULE_NAME)

OUTPUT_DIR = Path(os.environ.get("OUTPUT_DIR", Path(__file__).parent / "Output"))
OUTPUT_DIR.mkdir(exist_ok=True)

# ───────────────────────────────────────────────
# Utilitários internos
# ───────────────────────────────────────────────
def _read_e1_counts(xlsx_path: str) -> Tuple[int | None, int | None]:
    """Lê a célula E1 e devolve (expected, processed)."""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.worksheets[0]
        val = str(ws["E1"].value or "")
        m = re.search(r"(\d+)\s*/\s*(\d+)", val)
        if m:
            return int(m.group(1)), int(m.group(2))
    except Exception:
        pass
    return None, None


def _collect_debug_files(outdir: Path) -> List[str]:
    """Recolhe ficheiros de debug gerados pelo core."""
    debug_files = []
    for pattern in ["*_ocr_debug.txt", "*.csv", "process_summary_*.txt"]:
        for f in outdir.glob(pattern):
            debug_files.append(str(f))
    return debug_files


# ───────────────────────────────────────────────
# Função principal
# ───────────────────────────────────────────────
def process_pdf_with_stats(pdf_path: str) -> Tuple[List[str], Dict[str, Any], List[str]]:
    """
    Processa o PDF e devolve:
      - created_files: lista de ficheiros Excel criados
      - stats: resumo por requisição
      - debug_files: ficheiros auxiliares para /debug
    """
    rows_per_req = core.process_pdf_sync(pdf_path)
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    outdir = Path(os.environ.get("OUTPUT_DIR", OUTPUT_DIR))
    outdir.mkdir(exist_ok=True)

    created, per_req = [], []

    for i, rows in enumerate(rows_per_req, start=1):
        # 🔹 Filtrar entradas inválidas (strings, None, etc.)
        valid_rows = [r for r in rows if isinstance(r, dict) and "datarececao" in r]
        if not valid_rows:
            print(f"⚠️ Requisição {i} ignorada — sem amostras válidas.")
            continue

        fname = f"{base}.xlsx" if len(rows_per_req) == 1 else f"{base}_req{i}.xlsx"
        declared = valid_rows[0].get("declared_samples") if "declared_samples" in valid_rows[0] else None

        out_path = core.write_to_template(valid_rows, fname, expected_count=declared, source_pdf=pdf_path)
        if not out_path:
            continue
        created.append(out_path)

        expected, processed = _read_e1_counts(out_path)
        processed = processed or len(valid_rows)
        expected = expected or declared
        diff = processed - expected if expected is not None else None

        per_req.append({
            "req": i,
            "file": out_path,
            "processed": processed,
            "expected": expected,
            "diff": diff
        })

    stats = {
        "pdf_name": base,
        "req_count": len(per_req),
        "samples_total": sum(p["processed"] for p in per_req),
        "per_req": per_req,
    }

    debug_files = _collect_debug_files(outdir)
    return created, stats, debug_files


# ───────────────────────────────────────────────
# ZIP com summary e debug
# ───────────────────────────────────────────────
def build_zip_with_summary(excel_files: List[str], debug_files: List[str], summary_text: str) -> Tuple[bytes, str]:
    """Constrói um ZIP com ficheiros Excel, pasta debug e summary.txt."""
    mem = io.BytesIO()
    zip_name = f"xylella_output_{os.path.basename(os.getcwd())}_{os.getpid()}.zip"

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        # Excel na raiz
        for f in excel_files:
            if os.path.exists(f):
                z.write(f, arcname=os.path.basename(f))
        # pasta debug/
        for f in debug_files:
            if os.path.exists(f):
                z.write(f, arcname=f"debug/{os.path.basename(f)}")
        # summary.txt
        z.writestr("summary.txt", summary_text or "")

    mem.seek(0)
    return mem.read(), zip_name
