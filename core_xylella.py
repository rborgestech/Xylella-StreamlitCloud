# -*- coding: utf-8 -*-
"""
core_xylella.py — Cloud/Streamlit (OCR Azure direto + Parser Colab + Writer por requisição)

API exposta e usada pela UI (xylella_processor.py):
    • process_pdf_sync(pdf_path) -> List[str]   # devolve lista de paths dos Excels criados
    • process_folder_async(input_dir) -> str    # devolve path do ZIP criado
    • write_to_template(rows, out_name, expected_count=None, source_pdf=None) -> str  # escreve 1 XLSX com base no template

Requer:
  - AZURE_API_KEY, AZURE_ENDPOINT (env)
  - TEMPLATE_PATH (env) ou ficheiro 'TEMPLATE_PXf_SGSLABIP1056.xlsx' ao lado do core
  - OUTPUT_DIR (env) — diretório onde guardar .xlsx e _ocr_debug.txt (definido pela app por sessão)
"""

import os
import re
import time
import tempfile
import requests
import zipfile
import csv

from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, List, Optional

# 🟢 Biblioteca Excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from workalendar.europe import Portugal

# ───────────────────────────────────────────────
# Diretório base e template
# ───────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = Path(os.environ.get("TEMPLATE_PATH", BASE_DIR / "TEMPLATE_PXf_SGSLABIP1056.xlsx"))
if not TEMPLATE_PATH.exists():
    print(f"ℹ️ Aviso: TEMPLATE não encontrado em {TEMPLATE_PATH}. Será verificado no momento da escrita.")

# ───────────────────────────────────────────────
# Azure OCR — credenciais
# ───────────────────────────────────────────────
AZURE_API_KEY = os.environ.get("AZURE_API_KEY", "")
AZURE_ENDPOINT = os.environ.get("AZURE_ENDPOINT", "")
MODEL_ID = os.environ.get("AZURE_MODEL_ID", "prebuilt-document")

# ───────────────────────────────────────────────
# Estilos Excel
# ───────────────────────────────────────────────
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GRAY  = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
BOLD  = Font(bold=True, color="000000")
ITALIC= Font(italic=True, color="555555")

# ───────────────────────────────────────────────
# Cabeçalhos para distinguir templates
# ───────────────────────────────────────────────
HEADER_DGAV_PNPQ_RE = re.compile(
    r"Programa\s+nacional\s+de\s+Prospec[çc][aã]o\s+de\s+pragas\s+de\s+quarentena",
    re.I,
)

HEADER_ZONAS_DEM_RE = re.compile(
    r"Prospe[cç][aã]o\s+de\s*:\s*Xylella\s+fastidiosa\s+em\s+Zonas\s+Demarcadas",
    re.I,
)

HEADER_ICNF_NEW_RE = re.compile(
    r"ZONA\s+DEMARCADA\s*:", re.I
)

# ───────────────────────────────────────────────
# Utilitários genéricos
# ───────────────────────────────────────────────
def integrate_logic_and_generate_name(source_pdf: str) -> tuple[str, str]:
    """
    Função utilitária que:
    - Extrai a data (YYYYMMDD) do início do nome do ficheiro PDF
    - Adiciona 1 dia útil (PT)
    - Substitui esse prefixo de data pelo novo (YYYYMMDD) no nome final do ficheiro Excel

    Retorna:
        (data_ddmm, novo_nome_base)
    """
    base_name = os.path.splitext(os.path.basename(source_pdf))[0]

    m = re.match(r"(\d{8})_", base_name)
    if not m:
        return "0000", base_name

    try:
        data_envio = datetime.strptime(m.group(1), "%Y%m%d").date()
        cal = Portugal()
        data_util = cal.add_working_days(data_envio, 1)
        data_ddmm = data_util.strftime("%d%m")
        data_util_str = data_util.strftime("%Y%m%d")
    except Exception:
        return "0000", base_name

    novo_nome = re.sub(r"^\d{8}_", f"{data_util_str}_", base_name)
    return data_ddmm, novo_nome


def clean_value(s: str) -> str:
    if s is None:
        return ""
    if isinstance(s, (int, float)):
        return str(s)
    s = re.sub(r"[\u200b\t\r\f\v]+", " ", str(s))
    s = (s.strip()
           .replace("N/A", "")
           .replace("%", "")
           .replace("\n", " ")
           .replace("  ", " "))
    return s.strip()

# ───────────────────────────────────────────────
# Diretório de saída seguro — OBRIGATÓRIO
# (A app Streamlit define OUTPUT_DIR por sessão)
# ───────────────────────────────────────────────
def get_output_dir() -> Path:
    base = os.getenv("OUTPUT_DIR")
    if not base:
        raise RuntimeError(
            "OUTPUT_DIR não definido pela app. "
            "Defina os.environ['OUTPUT_DIR'] antes de usar o core_xylella."
        )

    d = Path(base)
    d.mkdir(parents=True, exist_ok=True)
    return d


def extract_all_text(result_json: Dict[str, Any]) -> str:
    """Concatena todo o texto linha a linha de todas as páginas."""
    lines = []
    for pg in result_json.get("analyzeResult", {}).get("pages", []):
        for ln in pg.get("lines", []):
            txt = (ln.get("content") or ln.get("text") or "").strip()
            if txt:
                lines.append(txt)
    return "\n".join(lines)

# ───────────────────────────────────────────────
# OCR Azure (PDF direto)
# ───────────────────────────────────────────────
def azure_analyze_pdf(pdf_path: str) -> Dict[str, Any]:
    if not AZURE_API_KEY or not AZURE_ENDPOINT:
        raise RuntimeError("Azure não configurado (AZURE_API_KEY/AZURE_ENDPOINT).")

    url = f"{AZURE_ENDPOINT.rstrip('/')}/formrecognizer/documentModels/{MODEL_ID}:analyze?api-version=2023-07-31"
    headers = {"Ocp-Apim-Subscription-Key": AZURE_API_KEY, "Content-Type": "application/pdf"}

    with open(pdf_path, "rb") as f:
        resp = requests.post(url, data=f.read(), headers=headers, timeout=120)
    if resp.status_code != 202:
        raise RuntimeError(f"Azure analyze falhou: {resp.status_code} {resp.text}")

    op = resp.headers.get("Operation-Location")
    if not op:
        raise RuntimeError("Azure não devolveu Operation-Location.")

    start = time.time()
    while True:
        r = requests.get(op, headers={"Ocp-Apim-Subscription-Key": AZURE_API_KEY}, timeout=60)
        j = r.json()
        st = j.get("status")
        if st == "succeeded":
            return j
        if st == "failed":
            raise RuntimeError(f"OCR Azure falhou: {j}")
        if time.time() - start > 180:
            raise RuntimeError("Timeout a aguardar OCR Azure.")
        time.sleep(1.2)

# ───────────────────────────────────────────────
# Parser — blocos do Colab (integrado)
# ───────────────────────────────────────────────
NATUREZA_KEYWORDS = [
    "ramos","folhas","ramosefolhas","ramosc/folhas","material","materialherbalho",
    "materialherbário","materialherbalo","natureza","insetos","sementes","solo"
]
TIPO_RE = re.compile(r"\b(Simples|Composta|Composto|Individual)\b", re.I)

def _looks_like_natureza(txt: str) -> bool:
    t = re.sub(r"\s+", "", (txt or "").lower())
    return any(k in t for k in NATUREZA_KEYWORDS)

def _clean_ref(raw: str) -> str:
    s = (raw or "").strip()
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"/{2,}", "/", s)
    s = re.sub(r"[A-Za-z]+", lambda m: m.group(0).upper(), s)
    s = s.replace("LUT", "LVT")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^A-Z0-9/]+$", "", s)
    return s

def detect_requisicoes(full_text: str):
    """Conta quantas requisições DGAV→SGS existem no texto OCR de um PDF."""
    pattern = re.compile(
        r"PROGRAMA\s+NACIONAL\s+DE\s+PROSPE[ÇC][AÃ]O\s+DE\s+PRAGAS\s+DE\s+QUARENTENA",
        re.IGNORECASE,
    )
    matches = list(pattern.finditer(full_text))
    count = len(matches)
    positions = [m.start() for m in matches]
    if count == 0:
        print("🔍 Nenhum cabeçalho encontrado — assumido 1 requisição.")
        count = 1
    else:
        print(f"🔍 Detetadas {count} requisições no ficheiro (posições: {positions})")
    return count, positions

def detect_document_type(full_text: str) -> str:
    """
    Decide se o PDF é ICNF ou DGAV com base nos cabeçalhos oficiais.
    - ICNF → 'Prospeção de: Xylella fastidiosa em Zonas Demarcadas'
    - DGAV → 'PROGRAMA NACIONAL DE PROSPEÇÃO DE PRAGAS DE QUARENTENA'
    """

    txt = full_text.upper()

    # 1) Formato novo ICNF – como o ficheiro que enviaste
    if "ENTIDADE:" in txt and "ICNF" in txt:
        return "ICNF"

    if "/XF/ICNFC" in txt:
        return "ICNF"

    # 2) Cabeçalho antigo ICNF com Xylella em zonas demarcadas
    if "XYLELLA FASTIDIOSA" in txt and "ZONA DEMARC" in txt and "PROSPEC" in txt:
        return "ICNF"

    # 3) DGAV – Programa Nacional
    if "PROGRAMA NACIONAL DE PROSPEC" in txt:
        return "DGAV"

    # 4) Fallback: referências ICNF sem DGAV
    if "/XF/ICNF" in txt and "DGAV" not in txt:
        return "ICNF"

    # 5) Fallback por omissão
    return "DGAV"

def split_if_multiple_requisicoes(full_text: str) -> List[str]:
    """Divide o texto OCR em blocos distintos, um por requisição DGAV→SGS."""
    text = full_text.replace("\r", "")
    text = re.sub(r"(\w)[\n\s]+(\w)", r"\1 \2", text)
    text = re.sub(r"(\d+)\s*/\s*([Xx][Ff])", r"\1/\2", text)
    text = re.sub(r"([Dd][Gg][Aa][Vv])[\s\n]*-", r"\1-", text)
    text = re.sub(r"([Ee][Dd][Mm])\s*/\s*(\d+)", r"\1/\2", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)

    pattern = re.compile(
        r"(?:PROGRAMA\s+NACIONAL\s+DE\s+PROSPE[ÇC][AÃ]O\s+DE\s+PRAGAS\s+DE\s+QUARENTENA)",
        re.IGNORECASE,
    )
    marks = [m.start() for m in pattern.finditer(text)]

    if not marks:
        print("🔍 Nenhum cabeçalho encontrado — tratado como 1 requisição.")
        return [text]
    if len(marks) == 1:
        print("🔍 Apenas 1 cabeçalho — 1 requisição detectada.")
        return [text]

    marks.append(len(text))
    blocos = []
    for i in range(len(marks) - 1):
        start = max(0, marks[i] - 200)
        end = min(len(text), marks[i + 1] + 200)
        bloco = text[start:end].strip()
        if len(bloco) > 400:
            blocos.append(bloco)
        else:
            print(f"⚠️ Bloco {i+1} demasiado pequeno ({len(bloco)} chars) — possivelmente OCR truncado.")
    print(f"🔍 Detetadas {len(blocos)} requisições distintas (por cabeçalho).")
    return blocos

def split_icnf_requisicoes(full_text: str) -> List[str]:
    """
    Divide o texto OCR em blocos distintos, um por requisição ICNF.
    Delimitador robusto: 'Zona demarcada:' (consistente em todos os ICNF novos).
    Mantém compatibilidade com o formato antigo.
    """
    text = full_text.replace("\r", "")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)

    # 1) Formato antigo ICNF (Prospeção Xylella)
    pattern_old = re.compile(
        r"Prospe[cç][aã]o\s+de\s*:\s*Xylella\s+fastidiosa\s+em\s+Zonas\s+Demarcadas",
        re.I,
    )
    marks = [m.start() for m in pattern_old.finditer(text)]

    # 2) Formato moderno ICNF – delimitador oficial
    if not marks:
        pattern_new = re.compile(
            r"ZONA\s+DEMARCADA\s*:?",   # apanha : ou ausência dele
            re.I,
        )
        marks = [m.start() for m in pattern_new.finditer(text)]

    if not marks:
        print("🔍 Nenhum cabeçalho ICNF encontrado — tratado como 1 requisição.")
        return [text]

    # Garantir que o último bloco é fechado
    marks.append(len(text))

    blocos: List[str] = []
    for i in range(len(marks) - 1):
        start = marks[i]
        end = marks[i + 1]
        bloco = text[start:end].strip()

        # Filtrar ruído (compatível com tua lógica anterior)
        if len(bloco) > 200:
            blocos.append(bloco)

    print(f"🟦 Detetadas {len(blocos)} requisições ICNF distintas.")
    return blocos or [text]

def normalize_date_str(val: str) -> str:
    """
    Corrige datas OCR partidas/coladas e devolve dd/mm/yyyy ou "".
    """
    if not val:
        return ""
    txt = str(val).strip().replace("-", "/").replace(".", "/")
    txt = re.sub(r"[\u00A0\s]+", "", txt)

    m_std = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", txt)
    if m_std:
        d, m_, y = map(int, m_std.groups())
        if 1 <= d <= 31 and 1 <= m_ <= 12 and 1900 <= y <= 2100:
            return f"{d:02d}/{m_:02d}/{y:04d}"

    if len(txt) >= 10 and txt[2] == "/" and txt[5] == "/":
        try:
            d, m_, y = int(txt[:2]), int(txt[3:5]), int(txt[6:10])
            if 1 <= d <= 31 and 1 <= m_ <= 12 and 1900 <= y <= 2100:
                return f"{d:02d}/{m_:02d}/{y:04d}"
        except Exception:
            pass

    digits = re.sub(r"\D", "", txt)

    if len(digits) == 8:
        d, m_, y = int(digits[:2]), int(digits[2:4]), int(digits[4:])
        if 1 <= d <= 31 and 1 <= m_ <= 12 and 1900 <= y <= 2100:
            return f"{d:02d}/{m_:02d}/{y:04d}"

    if len(digits) == 9:
        if digits[2:5] == "110":
            d, m_, y = int(digits[:2]), 10, int(digits[-4:])
            return f"{d:02d}/{m_:02d}/{y:04d}"
        d, m_, y = int(digits[:2]), int(digits[2:4]), int(digits[4:8])
        if 1 <= d <= 31 and 1 <= m_ <= 12:
            return f"{d:02d}/{m_:02d}/{y:04d}"

    m_flex = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", txt)
    if m_flex:
        d, m_, y = m_flex.groups()
        y = int(y) + (2000 if len(y) == 2 else 0)
        d, m_ = int(d), int(m_)
        if 1 <= d <= 31 and 1 <= m_ <= 12 and 1900 <= y <= 2100:
            return f"{d:02d}/{m_:02d}/{y:04d}"

    return ""

def _is_valid_date(value: str) -> bool:
    if isinstance(value, datetime):
        return True
    norm = normalize_date_str(value)
    if not norm:
        return False
    try:
        dt = datetime.strptime(norm, "%d/%m/%Y")
        return 1900 <= dt.year <= 2100
    except Exception:
        return False

def _to_datetime(value: str):
    if isinstance(value, datetime):
        return value
    norm = normalize_date_str(value)
    if not norm:
        return None
    try:
        dt = datetime.strptime(norm, "%d/%m/%Y")
        return dt if dt.year >= 1900 else None
    except Exception:
        return None

def extract_context_from_text(full_text: str):
    """
    Extrai informações gerais da requisição (zona, entidade DGAV/ICNF,
    datas (colheita/envio) e nº de amostras declaradas).
    """
    ctx: dict = {}

    # -----------------------------
    # Zona demarcada
    # -----------------------------
    m_zona = re.search(
        r"Zona\s+demarcada\s*:?\s*(.+?)(?=\s+Entidade\b|\s+T[ée]cnico\s+respons[aá]vel|\s+Data\s+de|\s+Datas?\s+de\s+recolha|$)",
        full_text,
        re.I | re.S,
    )
    if m_zona:
        zona = re.sub(r"\s+", " ", m_zona.group(1).strip())
        ctx["zona"] = zona
    else:
        m_old = re.search(r"Xylella\s+fastidiosa\s*\(([^)]+)\)", full_text, re.I)
        ctx["zona"] = m_old.group(1).strip() if m_old else "Zona Isenta"

    # -----------------------------
    # Entidade (limpa, sem ______, CAIXA X, etc.)
    # -----------------------------
    entidade = ""
    m_ent = re.search(r"Entidade\s*(?::|-)\s*(.+)", full_text, re.I)
    if m_ent:
        entidade = m_ent.group(1)
        entidade = entidade.split("\n")[0]              # só 1ª linha
        entidade = re.sub(r"[_\-–—]{2,}", " ", entidade)  # tira “______”, “-----”
        entidade = re.sub(r"CAIXA\s*\d+", "", entidade, flags=re.I)
        entidade = re.sub(r"\bCaixa\s*\d+\b", "", entidade, flags=re.I)
        entidade = re.sub(r"\s+", " ", entidade).strip()
        entidade = re.sub(r"[;,.\-]+$", "", entidade).strip()
    
        if entidade:
            entidade = re.sub(r"_+", "", entidade).strip()
    
    ctx["entidade"] = entidade

    # -----------------------------
    # Técnico responsável
    # -----------------------------
    tecnico = None
    m_tecnico = re.search(
        r"T[ée]cnico\s+respons[aá]vel\s*(?::|-)\s*(.+?)(?:\n|$|Data\s+(?:do|de)\s+envio|Data\s+(?:de\s+)?colheita|Datas?\s+de\s+recolha)",
        full_text,
        re.I | re.S,
    )
    if m_tecnico:
        tecnico = re.sub(r"(Data\s+.*)$", "", m_tecnico.group(1), flags=re.I).strip()
    ctx["responsavel_colheita"] = tecnico or ""

    # -----------------------------
    # DGAV (texto base para coluna "responsável amostra")
    # -----------------------------
    ctx["dgav"] = entidade or ""

    ctx["dgav"] = re.sub(
        r"T[ée]cnico\s+respons[aá]vel.*$",
        "",
        ctx["dgav"],
        flags=re.I
    ).strip()
    ctx["dgav"] = re.sub(r"respons[aá]vel$", "", ctx["dgav"], flags=re.I).strip()
    ctx["dgav"] = re.sub(r"[:;,.\-–—]+$", "", ctx["dgav"]).strip()

    # Fallback DGAV antigo
    if not ctx["dgav"]:
        responsavel_hdr, dgav = None, None
        m_hdr = re.search(
            r"Amostra(?:s|\(s\))?\s*colhida(?:s|\(s\))?\s*por\s*DGAV\s*[:\-]?\s*(.*)",
            full_text,
            re.IGNORECASE,
        )
        if m_hdr:
            tail = full_text[m_hdr.end():]
            linhas = [m_hdr.group(1)] + tail.splitlines()
            for ln in linhas[:4]:
                ln = (ln or "").strip()
                if ln:
                    responsavel_hdr = ln
                    break
            if responsavel_hdr:
                responsavel_hdr = re.sub(r"\S+@dgav\.pt|\S+@\S+", "", responsavel_hdr, flags=re.I)
                responsavel_hdr = re.sub(r"PROGRAMA.*|Data.*|N[º°].*", "", responsavel_hdr, flags=re.I)
                responsavel_hdr = re.sub(r"[:;,.\-–—]+$", "", responsavel_hdr).strip()

        if responsavel_hdr:
            if not re.match(r"^DGAV\b", responsavel_hdr, re.I):
                dgav = f"DGAV {responsavel_hdr}".strip()
            else:
                dgav = responsavel_hdr
        else:
            m_d = re.search(r"\bDGAV(?:\s+[A-Za-zÀ-ÿ?]+){1,4}", full_text)
            if m_d:
                dgav = re.sub(r"[:;,.\-–—]+$", "", m_d.group(0)).strip()

        ctx["dgav"] = dgav

    if ctx["dgav"] is None:
        ctx["dgav"] = ""

    # -----------------------------
    # Datas de colheita (robusto)
    # -----------------------------
    colheita_map: dict[str, str] = {}

    # Ex: "11/11/2025 (*)"
    for m in re.finditer(r"(\d{1,2}/\d{1,2}/\d{4})\s*\(\s*(\*+)\s*\)", full_text):
        colheita_map[f"({m.group(2).replace(' ', '')})"] = m.group(1)

    # 1) Tentativa clássica
    m_col = re.search(
        r"Datas?\s+de\s+recolha\s+de\s+amostras\s*[:\-\s]*([0-9/\-\s]+)",
        full_text,
        re.I,
    )

    default_colheita = ""
    # ICNF simples: "Data colheita das amostras: 3/11/2025"
    m_icnf_simple = re.search(
        r"Data\s+colheita\s+das?\s+amostras?\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{4})",
        full_text,
        re.I,
    )
    if m_icnf_simple:
        default_colheita = normalize_date_str(m_icnf_simple.group(1))

    if m_col:
        default_colheita = normalize_date_str(m_col.group(1))

    # 2) Reconstrução multi-linha (evitando linhas com "Total")
    if not default_colheita:
        m_block = re.search(
            r"Data\s+(?:de\s+)?colheita(?:\s+das?\s+amostras?)?\s*[:\-\s]*([\s\S]{0,60})",
            full_text,
            re.I,
        )
        if m_block:
            raw = m_block.group(1)
            raw = raw.replace("\n", " ").replace("\r", " ")
            digits = re.sub(r"[^\d]", "", raw)

            if len(digits) >= 8:
                candidate = f"{digits[:2]}/{digits[2:4]}/{digits[4:8]}"
                default_colheita = normalize_date_str(candidate) or ""


    # 3) Se existirem marcações (*), (**)
    if not colheita_map and default_colheita:
        for key in ("(*)", "(**)", "(***)"):
            colheita_map[key] = default_colheita

    ctx["colheita_map"] = colheita_map
    ctx["default_colheita"] = default_colheita

    # -----------------------------
    # Data de envio
    # -----------------------------
    m_envio = re.search(
        r"Data\s+(?:do|de)\s+envio(?:\s+das\s+amostras)?(?:\s+ao\s+laborat[oó]rio)?[:\-\s]*([0-9/\-\s]+)",
        full_text,
        re.I,
    )
    if not m_envio:
        m_envio = re.search(
            r"Data\s+envio\s+amostras?(?:\s+ao\s+laborat[oó]rio)?[:\-\s]*([0-9/\-\s]+)",
            full_text,
            re.I,
        )

    if m_envio:
        ctx["data_envio"] = normalize_date_str(m_envio.group(1))
    elif default_colheita:
        ctx["data_envio"] = default_colheita
    else:
        # fallback robusto → deixa vazio, o writer fará fallback seguro
        ctx["data_envio"] = ""

    # -----------------------------
    # Nº DE AMOSTRAS DECLARADAS — ultra robusto
    # -----------------------------
    lines = full_text.splitlines()
    flat  = re.sub(r"[ \t\r\n]+", " ", full_text)

    declared_samples = 0
    
    # 1) DGAV clássico (inclui ruído como "_ 2")
    m_dgav = re.search(
        r"N[º°o]?\s*de\s*amostras\s*neste\s*env[ií]o\s*[:\-]?\s*[_\-–—\.]*\s*(\d{1,3})",
        full_text,
        re.I,
    )
    if m_dgav:
        try:
            declared_samples = int(m_dgav.group(1))
        except:
            declared_samples = 0

    # 1) "Total: 27/35 amostras" → usa o MAIOR
    m = re.search(r"\bTotal\s*[:\-]?\s*(\d{1,3})\s*/\s*(\d{1,3})\s*amostras?", flat, re.I)
    if not m:
        m = re.search(r"\bTotal\s*[:\-]?\s*(\d{1,3})\s*/\s*(\d{1,3})\b", flat, re.I)
    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        if 0 < max(a, b) < 500:
            declared_samples = max(a, b)

    # 2) "Total: xx amostras 13"
    if declared_samples == 0:
        m = re.search(r"\bTotal\s*[:\-]?\s*[Xx]{1,3}\s*amostras?\s*(\d{1,3})\b", flat, re.I)
        if m:
            n = int(m.group(1))
            if 0 < n < 500:
                declared_samples = n

    # 3) "Total: 13 amostras" / "Total 13 amostras"
    if declared_samples == 0:
        m = re.search(r"\bTotal\s*[:\-]?\s*(\d{1,3})\s*amostras?\b", flat, re.I)
        if m:
            n = int(m.group(1))
            if 0 < n < 500:
                declared_samples = n

    # 4) "Total: 20"
    if declared_samples == 0:
        m = re.search(r"\bTotal\s*[:\-]?\s*(\d{1,3})\b", flat, re.I)
        if m:
            n = int(m.group(1))
            if 0 < n < 500:
                declared_samples = n

    # 5) Formato dividido:
    #       Total:
    #       13
    if declared_samples == 0:
        for i, ln in enumerate(lines):
            if re.match(r"^\s*Total\s*:?\s*$", ln.strip(), re.I):
                if i + 1 < len(lines):
                    nxt = re.sub(r"[^\d]", "", lines[i+1])
                    if nxt.isdigit():
                        n = int(nxt)
                        if 0 < n < 500:
                            declared_samples = n
                break

    # 6) Fallback DGAV clássico: "Nº de amostras: 2"
    if declared_samples == 0:
        m = re.search(r"\bN[º°o]?\s*de\s*amostras\s*[:\-]?\s*(\d{1,3})\b", flat, re.I)
        if m:
            n = int(m.group(1))
            if 0 < n < 500:
                declared_samples = n

    ctx["declared_samples"] = declared_samples
    print(f"📊 Nº de amostras declaradas detetadas (robusto): {declared_samples}")

    return ctx



def parse_xylella_tables(result_json, context, req_id=None) -> List[Dict[str, Any]]:
    """
    Extrai as amostras das tabelas Azure OCR para DGAV (Programa Nacional).
    """
    out: List[Dict[str, Any]] = []
    tables = result_json.get("analyzeResult", {}).get("tables", [])
    if not tables:
        print("⚠️ Nenhuma tabela encontrada.")
        return out

    for t in tables:
        nc = max(c.get("columnIndex", 0) for c in t.get("cells", [])) + 1
        nr = max(c.get("rowIndex", 0) for c in t.get("cells", [])) + 1
        grid = [[""] * nc for _ in range(nr)]
        for c in t.get("cells", []):
            grid[c["rowIndex"]][c["columnIndex"]] = clean_value(c.get("content", ""))

        for row in grid:
            if not row or not any(row):
                continue

            ref = _clean_ref(row[0]) if len(row) > 0 else ""
            if not ref or re.match(r"^\D+$", ref):
                continue

            hospedeiro = row[2] if len(row) > 2 else ""
            obs = row[3] if len(row) > 3 else ""

            if _looks_like_natureza(hospedeiro):
                hospedeiro = ""

            tipo = ""
            joined = " ".join([x for x in row if isinstance(x, str)])
            m_tipo = re.search(r"\b(Simples|Composta|Composto|Individual)\b", joined, re.I)
            if m_tipo:
                tipo = m_tipo.group(1).capitalize()
                if tipo.lower() == "composto":
                    tipo = "Composta"
                obs = re.sub(
                    r"\b(Simples|Composta|Composto|Individual)\b",
                    "",
                    obs,
                    flags=re.I,
                ).strip()

            datacolheita = context.get("default_colheita", "")
            m_ast = re.search(r"\(\s*\*+\s*\)", joined)
            if m_ast:
                mark = re.sub(r"\s+", "", m_ast.group(0))
                datacolheita = context.get("colheita_map", {}).get(mark, datacolheita)

            if obs.strip().lower() in ("simples", "composta", "composto", "individual"):
                obs = ""

            out.append({
                "requisicao_id": req_id,
                "datarececao": context.get("data_envio", ""),
                "datacolheita": datacolheita,
                "referencia": ref,
                "hospedeiro": hospedeiro,
                "tipo": tipo,
                "zona": context.get("zona", ""),
                "responsavelamostra": context.get("entidade") or context.get("dgav") or "",
                "responsavelcolheita": context.get("responsavel_colheita", ""),
                "observacoes": obs.strip(),
                "procedure": "XYLELLA",
                "datarequerido": context.get("data_envio", ""),
                "Score": "",
            })

    if not out:
        full_text = extract_all_text(result_json)
        pattern = re.compile(r"(\d{5,8}|[0-9]{1,3}/[A-Z]{1,3}/DGAV[-/]?\d{0,4})", re.I)
        matches = pattern.findall(full_text)
        if matches:
            for ref in matches:
                out.append({
                    "requisicao_id": req_id,
                    "datarececao": context.get("data_envio", ""),
                    "datacolheita": context.get("default_colheita", ""),
                    "referencia": ref.strip(),
                    "hospedeiro": "",
                    "tipo": "",
                    "zona": context.get("zona", ""),
                    "responsavelamostra": context.get("entidade") or context.get("dgav") or "",
                    "responsavelcolheita": context.get("responsavel_colheita", ""),
                    "observacoes": "",
                    "procedure": "XYLELLA",
                    "datarequerido": context.get("data_envio", ""),
                    "Score": "",
                })
            print(f"🔍 Fallback regex: {len(matches)} amostras detetadas.")

    print(f"✅ {len(out)} amostras extraídas no total (req_id={req_id}).")
    return out

def parse_icnf_zonas(full_text: str, ctx: dict, req_id: int = 1) -> List[Dict[str, Any]]:
    """
    Parser robusto para formulários de Zonas Demarcadas (DGAV ou ICNF).

    - Suporta:
        • linha com número + referência ("1 /XF/.....")
        • linha partida em duas ("1" + "/XF/....")
        • referência direta tipo "64/Xf/DGAVN/AMP/25"
        • evita cabeçalhos ("Refª da amostra", "Hospedeiro", "Tipo...")
        • extrai tipo, hospedeiro e referência corretamente
        • mapeia C3/C 3/C5/C 5 → "Composta"
    """

    # -----------------------------------------------
    # 1) LIMPAR E PRÉ-FILTRAR LINHAS
    # -----------------------------------------------
    lines = [l.strip() for l in full_text.splitlines() if l.strip()]

    header_garbage = (
        "refª", "refa", "refª da amostra",
        "hospedeiro",
        "tipo", "amostra simples", "amostra composta",
        "tipo (amostra simples", "composta)"
    )

    filtered = []
    for ln in lines:
        low = ln.lower()
        if any(h in low for h in header_garbage):
            continue
        filtered.append(ln)

    lines = filtered
    out: List[Dict[str, Any]] = []

    # -----------------------------------------------
    # 2) EXPRESSÕES REGULARES DE REFERÊNCIA / TIPO
    # -----------------------------------------------
    tipo_text_re = re.compile(r"\b(Simples|Composta|Composto|Individual)\b", re.I)
    # "1 /XF/..." ou "1 XF/..."
    ref_split_re = re.compile(r"^([1-9]\d{0,2})\s+(\/?XF\/[A-Z0-9\-/]+)", re.I)
    # "1/XF/..."
    ref_full_re = re.compile(r"^[1-9]\d{0,2}\s*/XF/[A-Z0-9\-/]+", re.I)
    # "64/Xf/..." (sem número de ordem)
    ref_direct_re = re.compile(r"^\d{1,3}\s*/?\s*[Xx][Ff]/[A-Z0-9\-/]+", re.I)

    # C3 / C 3 / C5 / C 5 → Composta
    tipo_c_re = re.compile(r"\bC\s*([35])\b", re.I)

    skip_if_no_ref = (
        "datas de recolha", "data de recolha", "data colheita",
        "total:", "total de amostras", "nº de amostras",
        "amostras"
    )

    pending_ref: Optional[str] = None
    pending_host: str = ""
    pending_tipo: str = ""

    # -----------------------------------------------
    # 3) FUNÇÃO PARA FECHAR UMA AMOSTRA
    # -----------------------------------------------
    def flush_sample(force: bool = False):
        nonlocal pending_ref, pending_host, pending_tipo
        if not pending_ref:
            return
        if not pending_host and not force:
            return

        tipo = pending_tipo or ""
        if tipo.lower() == "composto":
            tipo = "Composta"

        out.append({
            "requisicao_id": req_id,
            "datarececao": ctx.get("data_envio", ""),
            "datacolheita": ctx.get("default_colheita", ""),
            "referencia": pending_ref,
            "hospedeiro": pending_host.strip(),
            "tipo": tipo,
            "zona": ctx.get("zona", ""),
            "responsavelamostra": ctx.get("entidade", ""),
            "responsavelcolheita": ctx.get("responsavel_colheita", ""),
            "observacoes": "",
            "procedure": "XYLELLA",
            "datarequerido": ctx.get("data_envio", ""),
            "Score": "",
        })

        pending_ref = None
        pending_host = ""
        pending_tipo = ""

    # -----------------------------------------------
    # 4) LOOP PRINCIPAL DAS LINHAS
    # -----------------------------------------------
    i = 0
    while i < len(lines):
        ln = lines[i].strip()
        low = ln.lower()

        # Números soltos "1" + linha seguinte "/XF/..." (caso antigo)
        if re.fullmatch(r"[1-9]\d{0,2}", ln):
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                if nxt.upper().startswith(("/XF", "XF")):
                    ln = f"{ln} {nxt}"
                    lines[i + 1] = ""
                else:
                    i += 1
                    continue

        # 4.1 Referência direta "64/Xf/..."
        if ref_direct_re.match(ln):
            flush_sample(force=True)
            pending_ref = _clean_ref(ln)
            i += 1
            continue

        # 4.2 "1 /XF/..." com número de ordem + ref
        m_split = ref_split_re.match(ln)
        if m_split:
            flush_sample(force=True)
            num = m_split.group(1)
            ref = m_split.group(2)
            pending_ref = _clean_ref(f"{num} {ref}")
            i += 1
            continue

        # 4.3 "1/XF/..."
        if ref_full_re.match(ln):
            flush_sample(force=True)
            pending_ref = _clean_ref(ln)
            i += 1
            continue

        # Se ainda não temos referência, ignorar ruído
        if not pending_ref:
            if any(k in low for k in skip_if_no_ref):
                i += 1
                continue
            i += 1
            continue

        # Linhas de fecho de bloco ("Total...", "Data colheita...", etc.)
        # Não fechar bloco se a próxima linha começar com referência
        if any(k in low for k in skip_if_no_ref):
            if i + 1 < len(lines):
                nxt = lines[i+1].strip()
                # EXCEPÇÃO: se a próxima linha parecer uma referência → NÃO FECHAR
                if re.match(r"^\d{1,3}\s*/?\s*[Xx][Ff]/", nxt):
                    i += 1
                    continue
            flush_sample(force=True)
            i += 1
            continue


        # 4.4 Tipo textual (Simples / Composta / Individual)
        m_tipo_txt = tipo_text_re.search(ln)
        if m_tipo_txt:
            pending_tipo = m_tipo_txt.group(1).capitalize()
            host_part = ln[:m_tipo_txt.start()].strip()
            if host_part:
                pending_host = (pending_host + " " + host_part).strip() if pending_host else host_part
            flush_sample(force=True)
            i += 1
            continue

        # 4.5 Tipo "C 3" / "C3" / "C 5" / "C5" → Composta
        m_tipo_c = tipo_c_re.search(ln)
        if m_tipo_c:
            pending_tipo = "Composta"
            # tudo antes de "C 3"/"C 5" faz parte do hospedeiro (se existir)
            host_part = ln[:m_tipo_c.start()].strip()
            if host_part:
                pending_host = (pending_host + " " + host_part).strip() if pending_host else host_part
            flush_sample(force=True)
            i += 1
            continue

        # 4.6 Caso geral → parte do hospedeiro (pode vir em várias linhas)
        pending_host = (pending_host + " " + ln).strip() if pending_host else ln
        i += 1

    # -----------------------------------------------
    # 5) ÚLTIMA AMOSTRA
    # -----------------------------------------------
    flush_sample(force=False)

    print(f"🟦 parse_icnf_zonas: {len(out)} amostras extraídas (req {req_id})")
    return out

# ───────────────────────────────────────────────
# Dividir em requisições e extrair por bloco
# ───────────────────────────────────────────────
def parse_all_requisitions(result_json: Dict[str, Any], pdf_name: str, txt_path: str | None) -> List[Dict[str, Any]]:
    """
    Divide o documento em blocos (requisições) e devolve uma lista onde cada elemento
    é um dicionário: { "rows": [...amostras...], "expected": nº_declarado }.

    🔹 DGAV-PNPQ   → cabeçalho "Programa nacional de Prospecção de pragas de quarentena"
                     (pode ter 1 ou várias requisições no mesmo PDF, detetadas por esse cabeçalho)

    🔹 Zonas Demarcadas (DGAV ou ICNF) → cabeçalho
                     "Prospeção de: Xylella fastidiosa em Zonas Demarcadas"
                     (pode ter 1 ou várias requisições; cada cabeçalho = nova requisição)

    A ENTIDADE **NÃO** é usada para distinguir o modelo.
    O texto "CAIXA 1 / 2 / 3 / 4" é ignorado na lógica e apenas entra em `ctx["entidade"]`.
    """
    # Texto global OCR
    if txt_path and os.path.exists(txt_path):
        full_text = Path(txt_path).read_text(encoding="utf-8")
        print(f"📝 Contexto extraído de {os.path.basename(txt_path)}")
    else:
        full_text = extract_all_text(result_json)

    # ------------------------------------------------------------
    # 1) Detetar template pelo cabeçalho (NUNCA pela 'Entidade')
    # ------------------------------------------------------------
    is_dgav_pnpq = bool(HEADER_DGAV_PNPQ_RE.search(full_text))
    is_zonas_dem = (
    bool(HEADER_ZONAS_DEM_RE.search(full_text))      # ICNF antigo
    or bool(HEADER_ICNF_NEW_RE.search(full_text))    # ICNF moderno
    )

    # ------------------------------------------------------------
    # 🟦 ZONAS DEMARCADAS (DGAV ou ICNF) — parser de linhas
    # ------------------------------------------------------------
    if is_zonas_dem and not is_dgav_pnpq:
        print("🟦 Documento 'Zonas Demarcadas' detetado — parser exclusivo de linhas ativado.")
        blocos = split_icnf_requisicoes(full_text)
        if not blocos:
            blocos = [full_text]

        results: List[Dict[str, Any]] = []
        for i, bloco in enumerate(blocos, start=1):
            ctx = extract_context_from_text(bloco)
            rows = parse_icnf_zonas(bloco, ctx, req_id=i)
            expected = ctx.get("declared_samples", len(rows))

            # Segurança extra: se extraiu mais linhas do que o declarado, corta ao declarado
            #if expected and len(rows) > expected:
            #    print(
            #        f"⚠️ Zonas Demarcadas bloco {i}: {len(rows)} amostras extraídas > declaradas {expected}. "
            #        "Cortar para o nº declarado."
            #    )
            #    rows = rows[:expected]

            results.append({"rows": rows, "expected": expected})

        return results

    # ------------------------------------------------------------
    # 🟧 DGAV PNPQ (Programa nacional de Prospecção de pragas de quarentena)
    #     – lógica original baseada em tabelas Azure
    # ------------------------------------------------------------
    print("🟧 Documento tratado como DGAV PNPQ (Programa nacional).")
    count, _ = detect_requisicoes(full_text)
    all_tables = result_json.get("analyzeResult", {}).get("tables", []) or []

    # Caso simples (1 requisição DGAV)
    if count <= 1:
        context = extract_context_from_text(full_text)
        amostras = parse_xylella_tables(result_json, context, req_id=1)
        expected = context.get("declared_samples", len(amostras))
        return [{"rows": amostras, "expected": expected}]

    # Múltiplas requisições DGAV — segmentar por cabeçalhos
    blocos = split_if_multiple_requisicoes(full_text)
    num_blocos = len(blocos)
    out: List[List[Dict[str, Any]]] = [[] for _ in range(num_blocos)]

    # Extrair referências por bloco
    refs_por_bloco: List[List[str]] = []
    for i, bloco in enumerate(blocos, start=1):
        refs_bloco = re.findall(
            r"\b\d{1,3}/[A-Z]{0,2}/DGAV(?:-[A-Z0-9/]+)?|\b\d{2,4}/\d{2,4}/[A-Z0-9\-]+",
            bloco, re.I
        )
        refs_bloco = [r.strip() for r in refs_bloco if len(r.strip()) > 4]
        print(f"   ↳ Bloco {i}: {len(refs_bloco)} referências detectadas")
        refs_por_bloco.append(refs_bloco)

    # Texto de cada tabela
    table_texts = [
        " ".join(c.get("content", "") for c in t.get("cells", []))
        for t in all_tables
    ]

    # Atribuição exclusiva de tabelas por bloco
    assigned_to: List[int] = [-1] * len(all_tables)
    for ti, ttxt in enumerate(table_texts):
        scores = []
        for bi, refs in enumerate(refs_por_bloco):
            if not refs:
                scores.append(0)
                continue
            cnt = sum(1 for r in refs if r in ttxt)
            scores.append(cnt)
        best = max(scores) if scores else 0
        if best > 0:
            bi = scores.index(best)
            assigned_to[ti] = bi

    # fallback: tabelas não atribuídas → distribuição uniforme
    unassigned = [i for i, b in enumerate(assigned_to) if b < 0]
    if unassigned:
        for k, ti in enumerate(unassigned):
            assigned_to[ti] = k % num_blocos

    # Construir amostras por bloco com base na atribuição
    for bi in range(num_blocos):
        try:
            context = extract_context_from_text(blocos[bi])
            tables_filtradas = [
                all_tables[ti]
                for ti in range(len(all_tables))
                if assigned_to[ti] == bi
            ]
            if not tables_filtradas:
                print(f"⚠️ Bloco {bi+1}: sem tabelas atribuídas (usar todas como fallback).")
                tables_filtradas = all_tables

            local = {"analyzeResult": {"tables": tables_filtradas}}
            amostras = parse_xylella_tables(local, context, req_id=bi+1)
            out[bi] = amostras or []
        except Exception as e:
            print(f"❌ Erro no bloco {bi+1}: {e}")
            out[bi] = []

    out = [req for req in out if req]
    print(f"\n🏁 Concluído: {len(out)} requisições com amostras extraídas (atribuição exclusiva).")

    results: List[Dict[str, Any]] = []
    for bi, bloco in enumerate(blocos[:len(out)], start=1):
        ctx = extract_context_from_text(bloco)
        expected = ctx.get("declared_samples", 0)
        results.append({
            "rows": out[bi - 1],
            "expected": expected
        })
    return results



# ───────────────────────────────────────────────
# Datas úteis e nomes de ficheiro
# ───────────────────────────────────────────────
def get_next_business_day(date_str: str) -> str:
    """
    Recebe string de data (dd/mm/yyyy ou yyyymmdd) e devolve próximo dia útil em formato YYYYMMDD.
    """
    if not date_str:
        return datetime.now().strftime("%Y%m%d")

    s = str(date_str).strip()
    try:
        if re.match(r"^\d{8}$", s):
            dt = datetime.strptime(s, "%Y%m%d").date()
        else:
            norm = normalize_date_str(s)
            dt = datetime.strptime(norm, "%d/%m/%Y").date()
    except Exception:
        return datetime.now().strftime("%Y%m%d")

    cal = Portugal()
    next_bd = cal.add_working_days(dt, 1)
    return next_bd.strftime("%Y%m%d")

def gerar_nome_excel_corrigido(source_pdf: str, data_envio: str) -> str:
    """
    Substitui a data inicial do nome do PDF pela nova data (com +1 útil).
    Ex: 20251030_ReqX19_27-10 Formulário.pdf -> 20251031_ReqX19_27-10 Formulário.xlsx
    """
    base_pdf = Path(source_pdf).name
    nova_data = get_next_business_day(data_envio)  # YYYYMMDD
    nome_corrigido = re.sub(r"^\d{8}_", f"{nova_data}_", base_pdf)
    return nome_corrigido.replace(".pdf", ".xlsx")

# ───────────────────────────────────────────────
# Escrita no TEMPLATE — 1 ficheiro por requisição
# ───────────────────────────────────────────────
def write_to_template(ocr_rows, out_name, expected_count=None, source_pdf=None):
    if not ocr_rows:
        print(f"⚠️ {out_name}: sem linhas para escrever.")
        return None

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template não encontrado: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.worksheets[0]
    start_row = 4

    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill   = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    bold_center = Font(bold=True, color="000000")

    for row in range(start_row, 201):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
        ws[f"I{row}"].value = None

    def normalize_date_str_local(val: str) -> str:
        if not val:
            return ""
        s = re.sub(r"\D", "", str(val))
        if len(s) >= 8:
            d, m, y = int(s[:2]), int(s[2:4]), int(s[4:8])
            if 1 <= d <= 31 and 1 <= m <= 12:
                return f"{d:02d}/{m:02d}/{y:04d}"
        m = re.match(r"^\s*(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})\s*$", str(val))
        if m:
            d, m_, y = map(int, m.groups())
            return f"{d:02d}/{m_:02d}/{y:04d}"
        return str(val).strip()

    def to_excel_date(val: str):
        s = normalize_date_str_local(val)
        try:
            return datetime.strptime(s, "%d/%m/%Y")
        except Exception:
            return None

    base = Path(source_pdf or out_name).name
    m = re.search(r"(X\d{2,3})", base, flags=re.I)
    req_id = m.group(1).upper() if m else "X??"

    last_next_bd = None

    for idx, row in enumerate(ocr_rows, start=start_row):
        rececao_val = row.get("datarececao", "").strip()

        # Se o parser não forneceu uma data válida → usar fallback seguro
        if not normalize_date_str_local(rececao_val):
            # extrair data do nome do PDF
            base_pdf = Path(source_pdf).stem
            mdate = re.match(r"(\d{8})_", base_pdf)
            if mdate:
                ymd = mdate.group(1)
                try:
                    dt_tmp = datetime.strptime(ymd, "%Y%m%d").date()
                    cal = Portugal()
                    rececao_dt = cal.add_working_days(dt_tmp, 1)
                    rececao_val = rececao_dt.strftime("%d/%m/%Y")
                except:
                    pass

        colheita_val = row.get("datacolheita", "")

        base_date = normalize_date_str_local(rececao_val)
        if base_date and re.match(r"\d{2}/\d{2}/\d{4}", str(base_date)):
            try:
                cal = Portugal()
                dt = datetime.strptime(base_date, "%d/%m/%Y").date()
                next_bd = cal.add_working_days(dt, 1)
                last_next_bd = next_bd
                ws[f"A{idx}"].value = next_bd
                ws[f"A{idx}"].number_format = "dd/mm/yyyy"
                ws[f"L{idx}"].value = f"=A{idx}+30"
                ws[f"L{idx}"].number_format = "dd/mm/yyyy"
            except Exception:
                ws[f"A{idx}"].value = base_date
                ws[f"A{idx}"].fill = red_fill
                ws[f"L{idx}"].value = ""
                ws[f"L{idx}"].fill = red_fill
        else:
            ws[f"A{idx}"].value = str(rececao_val or "").strip()
            ws[f"A{idx}"].fill = red_fill
            ws[f"L{idx}"].value = ""
            ws[f"L{idx}"].fill = red_fill

        cell_B = ws[f"B{idx}"]
        dt_colheita = to_excel_date(colheita_val)
        if dt_colheita:
            cell_B.value = dt_colheita
            cell_B.number_format = "dd/mm/yyyy"
        else:
            norm = normalize_date_str_local(colheita_val)
            cell_B.value = norm or str(colheita_val).strip()
            cell_B.fill = red_fill

        ws[f"C{idx}"] = row.get("referencia", "")
        ws[f"D{idx}"] = row.get("hospedeiro", "")
        ws[f"E{idx}"] = row.get("tipo", "")
        ws[f"F{idx}"] = row.get("zona", "")
        ws[f"G{idx}"] = row.get("responsavelamostra", "")
        ws[f"H{idx}"] = row.get("responsavelcolheita", "")
        ws[f"I{idx}"] = ""

        ws[f"J{idx}"] = f'=TEXT(A{idx},"ddmm")&"{req_id}."&TEXT(ROW()-3,"000")'
        ws[f"K{idx}"] = row.get("procedure", "")

        ws[f"L{idx}"].value = f"=A{idx}+30"
        ws[f"L{idx}"].number_format = "dd/mm/yyyy"

        for col in ("A", "B", "C", "D", "E", "F", "G"):
            c = ws[f"{col}{idx}"]
            if not c.value or str(c.value).strip() == "":
                c.fill = red_fill

        if row.get("WasCorrected") or row.get("ValidationStatus") in ("review", "unknown", "no_list"):
            ws[f"D{idx}"].fill = yellow_fill

    processed = len(ocr_rows)
    expected = expected_count
    ws.merge_cells("E1:F1")
    cell = ws["E1"]
    val_str = f" {expected or 0} / {processed}"
    cell.value = f"Nº Amostras (Dec./Proc.): {val_str}"
    cell.font = bold_center
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = red_fill if (expected is not None and expected != processed) else green_fill

    ws.merge_cells("G1:J1")
    pdf_orig_name = Path(source_pdf).name if source_pdf else "(desconhecida)"
    ws["G1"].value = f"Origem: {pdf_orig_name}"
    ws["G1"].font = Font(italic=True, color="555555")
    ws["G1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["G1"].fill = gray_fill
    # ───────────────────────────────────────────────
    # 🕒 Data/hora do processamento (Excel)
    # ───────────────────────────────────────────────
    ws.merge_cells("K1:L1")
    ws["K1"].value = f"Processado em: {datetime.now():%d/%m/%Y %H:%M}"
    ws["K1"].font = Font(italic=True, color="333333")
    ws["K1"].alignment = Alignment(horizontal="right", vertical="center")
    ws["K1"].fill = gray_fill

    if last_next_bd:
        data_envio = last_next_bd
    else:
        data_envio = datetime.now().date()

    if not isinstance(data_envio, datetime):
        data_envio = datetime.combine(data_envio, datetime.min.time())

    data_util = data_envio.strftime("%Y%m%d")

    base_name = Path(out_name).stem
    base_name = re.sub(r"^\d{8}_", "", base_name)

    new_name = f"{data_util}_{base_name}.xlsx"

    out_path = get_output_dir() / new_name
    wb.save(out_path)

    print(f"📁 Ficheiro gravado: {out_path}")
    return str(out_path)

# ───────────────────────────────────────────────
# Log opcional (compatível com o teu Colab)
# ───────────────────────────────────────────────
def append_process_log(pdf_name, req_id, processed, expected, out_path=None, status="OK", error_msg=None):
    out_dir = get_output_dir()
    log_path = out_dir / "process_log.csv"
    today_str = datetime.now().strftime("%Y-%m-%d")
    summary_path = out_dir / f"process_summary_{today_str}.txt"

    exists = os.path.exists(log_path)
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        if not exists:
            writer.writerow(["DataHora","PDF","ReqID","Processadas","Requisitadas","OutputExcel","Status","Mensagem"])
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        writer.writerow([ts, os.path.basename(pdf_name), req_id, processed, expected or "", out_path or "", status, error_msg or ""])

    try:
        with open(summary_path, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {os.path.basename(pdf_name)} | Req {req_id} | {processed}/{expected or '?'} | {status} {os.path.basename(out_path or '')}\n")
    except Exception:
        pass

# ───────────────────────────────────────────────
# API pública usada pela app Streamlit
# ───────────────────────────────────────────────
def process_pdf_sync(pdf_path: str) -> list[str]:
    """
    Processa um único PDF:
      - executa OCR Azure,
      - extrai requisições e amostras,
      - gera 1 ficheiro Excel por requisição.
    Retorna: lista de caminhos absolutos dos ficheiros Excel criados.
    """
    base = os.path.basename(pdf_path)
    print(f"\n🧪 Início de processamento: {base}")

    result_json = azure_analyze_pdf(pdf_path)

    txt_path = get_output_dir() / f"{Path(base).stem}_ocr_debug.txt"
    txt_path.write_text(extract_all_text(result_json), encoding="utf-8")
    print(f"📝 Texto OCR bruto guardado em: {txt_path}")

    req_results = parse_all_requisitions(result_json, pdf_path, str(txt_path))

    valid_reqs = [req for req in req_results if req.get("rows")]
    total_amostras = sum(len(req["rows"]) for req in valid_reqs)
    print(f"✅ {base}: {len(valid_reqs)} requisição(ões) válidas, {total_amostras} amostras extraídas.")

    created_files = []
    for i, req in enumerate(valid_reqs, start=1):
        rows = req.get("rows", [])
        expected = req.get("expected", 0)

        if not rows:
            continue

        base_name = Path(pdf_path).stem
        out_name = f"{base_name}_req{i}.xlsx" if len(valid_reqs) > 1 else f"{base_name}.xlsx"

        out_path = write_to_template(rows, out_name, expected_count=expected, source_pdf=pdf_path)
        created_files.append(out_path)
        print(f"💾 Excel criado: {out_path}")

    print(f"🏁 {base}: {len(created_files)} ficheiro(s) Excel gerado(s).")
    return [str(f) for f in created_files if Path(f).exists()]

# ───────────────────────────────────────────────
# Processamento em lote (pasta)
# ───────────────────────────────────────────────
def process_folder_async(input_dir: str) -> str:
    """
    Processa todos os PDFs em `input_dir` chamando `process_pdf_sync(pdf_path)`.
    Usa SEMPRE o OUTPUT_DIR da sessão (definido pelo app.py).
    Cria ZIP final com:
      • todos os XLSX gerados
      • summary.txt
    Retorna o caminho completo do ZIP criado dentro do OUTPUT_DIR da sessão.
    """
    out_dir = get_output_dir()
    out_dir.mkdir(parents=True, exist_ok=True)

    start_time = time.time()
    input_path = Path(input_dir)
    pdf_files = sorted(input_path.glob("*.pdf"))

    if not pdf_files:
        print("⚠️ Nenhum PDF encontrado na pasta.")
        return ""

    print(f"📂 Início do processamento: {input_path} ({len(pdf_files)} PDF(s))")

    all_excels = []

    for pdf_path in pdf_files:
        base = pdf_path.name
        print(f"\n🔹 A processar: {base}")

        try:
            created = process_pdf_sync(str(pdf_path))
            excels = [f for f in created if f.lower().endswith(".xlsx")]
            all_excels.extend(excels)
            print(f"✅ {base}: {len(excels)} ficheiro(s) Excel.")
        except Exception as e:
            print(f"❌ Erro ao processar {base}: {e}")

    elapsed_time = time.time() - start_time

    summary_path = out_dir / "summary.txt"
    with open(summary_path, "w", encoding="utf-8") as f:
        for pdf_path in pdf_files:
            base = pdf_path.name
            related_excels = [e for e in all_excels if Path(base).stem in Path(e).stem]
            f.write(f"{base}: {len(related_excels)} requisição(ões)\n")
            for e in related_excels:
                f.write(f"   ↳ {Path(e).name}\n")
            f.write("\n")

        f.write("───────────────────────────\n")
        f.write(f"📊 Total de ficheiros Excel: {len(all_excels)}\n")
        f.write(f"⏱️ Tempo total: {elapsed_time:.1f} segundos\n")
        f.write(f"📅 Executado em: {datetime.now():%d/%m/%Y às %H:%M:%S}\n")

    print(f"🧾 Summary criado: {summary_path}")

    base_name = Path(pdf_files[0]).stem
    zip_name = f"{base_name}_output.zip"
    zip_path = out_dir / zip_name

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for e in all_excels:
            e_path = Path(e)
            if e_path.exists():
                zipf.write(e_path, e_path.name)

        if summary_path.exists():
            zipf.write(summary_path, summary_path.name)

    print(f"📦 ZIP final criado: {zip_path}")
    print(f"✅ Processamento completo ({elapsed_time:.1f}s).")

    return str(zip_path)



































