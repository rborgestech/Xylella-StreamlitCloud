# -*- coding: utf-8 -*-
import io
import zipfile
from datetime import datetime
from typing import List, Tuple

import streamlit as st
from openpyxl import load_workbook

from processor import process_pre_to_dgav, REQUIRED_DGAV_COLS


# ───────────────────────────────────────────────
# Funções auxiliares
# ───────────────────────────────────────────────
def analyse_output_xlsx(xlsx_bytes: bytes) -> Tuple[int, List[str]]:
    """
    Analisa o ficheiro DGAV gerado:
      - Conta o nº de amostras (linhas com CODIGO_AMOSTRA)
      - Devolve lista de warnings (colunas obrigatórias com células vazias)
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Default"]

    # Mapear cabeçalhos
    header_indices = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            header_indices[str(val)] = col

    warnings = []
    # nº de amostras = nº de linhas com CODIGO_AMOSTRA não vazio
    codigo_col_idx = header_indices.get("CODIGO_AMOSTRA")
    sample_count = 0
    if codigo_col_idx:
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=codigo_col_idx).value
            if v not in (None, ""):
                sample_count += 1

    # analisar colunas obrigatórias
    for col_name in REQUIRED_DGAV_COLS:
        col_idx = header_indices.get(col_name)
        if col_idx is None:
            warnings.append(f"Coluna obrigatória ausente no output: {col_name}")
            continue

        any_value = False
        any_empty = False

        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v not in (None, ""):
                any_value = True
            else:
                any_empty = True

        if not any_value:
            warnings.append(f"Coluna obrigatória sem registos: {col_name}")
        elif any_empty:
            warnings.append(f"Coluna obrigatória com células vazias: {col_name}")

    return sample_count, warnings


def build_zip_with_summary(
    outputs: List[Tuple[str, bytes, int, List[str]]],
    summary_lines: List[str],
    timestamp: str,
) -> bytes:
    """
    Cria um ZIP em memória com:
      - 1 ficheiro DGAV por input
      - summary.txt com o resumo do processamento
    """
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        # ficheiros Excel
        for original_name, data, sample_count, warnings in outputs:
            base = original_name.rsplit(".", 1)[0]
            out_name = f"{base}_DGAV_{timestamp}.xlsx"
            z.writestr(out_name, data)

        # summary.txt
        z.writestr("summary.txt", "\n".join(summary_lines))

    mem.seek(0)
    return mem.getvalue()


# ───────────────────────────────────────────────
# Configuração base
# ───────────────────────────────────────────────
st.set_page_config(page_title="Xylella → DGAV", page_icon="🧪", layout="centered")

st.title("🧪 Xylella – Conversor Pré-registo → DGAV")
st.caption("Carrega ficheiros 'AVALIAÇÃO PRÉ-REGISTO – Amostras Xylella' e gera automaticamente o ficheiro DGAV.")

# CSS semelhante ao processor de PDFs
st.markdown(
    """
<style>
.stButton > button[kind="primary"]{
  background:#CA4300!important;border:1px solid #CA4300!important;color:#fff!important;
  font-weight:600!important;border-radius:6px!important;transition:background-color .2s ease-in-out!important;
}
.stButton > button[kind="primary"]:hover{background:#A13700!important;border-color:#A13700!important;}
[data-testid="stFileUploader"]>div:first-child{
  border:2px dashed #CA4300!important;border-radius:10px!important;padding:1rem!important
}

/* Caixas de estado */
.file-box{border-radius:8px;padding:.6rem 1rem;margin-bottom:.5rem;opacity:0;
          animation:fadeIn .4s ease forwards}
@keyframes fadeIn{from{opacity:0;transform:translateY(-4px)}to{opacity:1;transform:translateY(0)}}
.file-box.success{background:#e6f9ee;border-left:4px solid #1a7f37}
.file-box.warning{background:#fff8e5;border-left:4px solid #e6a100}
.file-box.error{background:#fdeaea;border-left:4px solid #cc0000}
.file-title{font-size:.9rem;font-weight:600;color:#1A365D}
.file-sub{font-size:.8rem;color:#2A4365}

/* Pontinhos animados */
.dots::after{content:'...';display:inline-block;animation:dots 1.5s steps(4,end) infinite}
@keyframes dots{
  0%,20%{color:rgba(42,67,101,0);text-shadow:.25em 0 0 rgba(42,67,101,0),.5em 0 0 rgba(42,67,101,0)}
  40%{color:#2A4365;text-shadow:.25em 0 0 rgba(42,67,101,0),.5em 0 0 rgba(42,67,101,0)}
  60%{text-shadow:.25em 0 0 #2A4365,.5em 0 0 rgba(42,67,101,0)}
  80%,100%{text-shadow:.25em 0 0 #2A4365,.5em 0 0 #2A4365}
}
</style>
""",
    unsafe_allow_html=True,
)

# ───────────────────────────────────────────────
# Upload de ficheiros
# ───────────────────────────────────────────────
uploads = st.file_uploader(
    "📂 Carrega um ou vários ficheiros de pré-registo (XLSX)",
    type=["xlsx"],
    accept_multiple_files=True,
)

process_btn = st.button("📄 Processar ficheiros de Input", type="primary") if uploads else None

if not uploads:
    st.info("💡 Carrega pelo menos um ficheiro de pré-registo para ativar o processamento.")

# ───────────────────────────────────────────────
# Processamento
# ───────────────────────────────────────────────
if uploads and process_btn:
    total = len(uploads)
    progress = st.progress(0.0)

    outputs: List[Tuple[str, bytes, int, List[str]]] = []
    summary_lines: List[str] = []
    total_samples = 0
    warning_files = 0
    error_files = 0

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for i, up in enumerate(uploads, start=1):
        placeholder = st.empty()
        placeholder.markdown(
            f"""
            <div class='file-box'>
              <div class='file-title'>📄 {up.name}</div>
              <div class='file-sub'>Ficheiro {i} de {total} — a processar<span class="dots"></span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        try:
            # Garantir leitura em memória apenas
            data_in = io.BytesIO(up.getbuffer())

            # Converte pré-registo → DGAV (em memória)
            output_bytes, log_msg = process_pre_to_dgav(data_in)

            # Analisa ficheiro DGAV para nº de amostras e warnings de colunas
            sample_count, col_warnings = analyse_output_xlsx(output_bytes)
            total_samples += sample_count

            status_class = "success"
            extra_html = ""
            summary_line = f"{up.name}: {sample_count} amostra(s). {log_msg}"

            if col_warnings:
                status_class = "warning"
                warning_files += 1
                bullets = "<br>".join(f"• {w}" for w in col_warnings)
                extra_html = f"<div class='file-sub'>⚠️ Colunas vazias/irregulares:<br>{bullets}</div>"
                summary_line += " ⚠ " + " | ".join(col_warnings)

            outputs.append((up.name, output_bytes, sample_count, col_warnings))

            html = (
                f"<div class='file-box {status_class}'>"
                f"<div class='file-title'>📄 {up.name}</div>"
                f"<div class='file-sub'><b>{sample_count}</b> amostra(s) processadas.</div>"
                f"{extra_html}</div>"
            )
            placeholder.markdown(html, unsafe_allow_html=True)
            summary_lines.append(summary_line)

        except Exception as e:
            error_files += 1
            err_msg = f"{up.name}: erro ao processar ({e})"
            summary_lines.append(err_msg)
            html = (
                f"<div class='file-box error'>"
                f"<div class='file-title'>📄 {up.name}</div>"
                f"<div class='file-sub'>❌ Erro ao processar: {e}</div>"
                f"</div>"
            )
            placeholder.markdown(html, unsafe_allow_html=True)

        progress.progress(i / total)

    # ───────────────────────────────────────────
    # Resumo e downloads
    # ───────────────────────────────────────────
    st.markdown(
        f"""
        <div style='text-align:center;margin-top:1.5rem;'>
          <h3>🏁 Processamento concluído!</h3>
          <p>Foram processados <b>{len(outputs)}</b> ficheiro(s) válido(s),
          com um total de <b>{total_samples}</b> amostras.<br>
          Ficheiros com avisos (colunas vazias): <b>{warning_files}</b>.<br>
          Ficheiros com erro: <b>{error_files}</b>.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Acrescentar resumo global
    summary_lines.append("")
    summary_lines.append(f"Total de ficheiros válidos: {len(outputs)}")
    summary_lines.append(f"Total de amostras: {total_samples}")
    summary_lines.append(f"Ficheiros com avisos (colunas vazias): {warning_files}")
    summary_lines.append(f"Ficheiros com erro: {error_files}")
    summary_lines.append(f"Executado em: {datetime.now():%d/%m/%Y às %H:%M:%S}")

    if len(outputs) == 1:
        # download direto do único ficheiro DGAV
        original_name, data, sample_count, col_warnings = outputs[0]
        base = original_name.rsplit(".", 1)[0]
        out_name = f"{base}_DGAV_{timestamp}.xlsx"

        st.download_button(
            label="⬇️ Descarregar ficheiro DGAV",
            data=data,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # também disponibilizar summary.txt isolado se quiseres
        st.download_button(
            label="📝 Descarregar resumo (summary.txt)",
            data="\n".join(summary_lines),
            file_name=f"summary_{timestamp}.txt",
            mime="text/plain",
        )
    elif len(outputs) > 1:
        # criar ZIP em memória
        zip_bytes = build_zip_with_summary(outputs, summary_lines, timestamp)
        zip_name = f"xylella_dgav_{timestamp}.zip"

        st.download_button(
            label="📦 Descarregar resultados (ZIP)",
            data=zip_bytes,
            file_name=zip_name,
            mime="application/zip",
        )
